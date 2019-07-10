# !/usr/bin/env python
# coding:gbk
import sys, re, os, time, random, types, traceback, datetime, smtplib, openpyxl
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from email.mime.text import MIMEText
from email.header import Header
from openpyxl import load_workbook


def sleeptime(hour, min, sec):
    return hour * 3600 + min * 60 + sec


def smtp(zw1, fs1, js1, zt1, stmpset):
    mail_host = "smtp.exmail.qq.com"
    mail_user = stmpset[0]
    mail_pass = stmpset[1]

    sender = stmpset[2]
    receivers = stmpset[3]

    message = MIMEText(zw1, 'plain', 'gbk')
    message['From'] = Header(fs1, 'gbk')
    message['To'] = Header(js1, 'gbk')

    subject = zt1
    message['Subject'] = Header(subject, 'gbk')

    try:
        smtpobj = smtplib.SMTP()
        smtpobj.connect(mail_host, 25)    # 25 为 SMTP 端口号
        smtpobj.login(mail_user, mail_pass)
        smtpobj.sendmail(sender, receivers, message.as_string())
        print("邮件发送成功")
    except smtplib.SMTPException:
        print("Error: 无法发送邮件")


def get_data():
    today = time.strftime('%Y-%m-%d', time.localtime())
    path = os.path.dirname(sys.argv[0])
    os.chdir(path)
    wb = load_workbook(u'付款审核.xlsx', read_only=True, data_only=True)

    ws = wb[u'payout']
    rows_len = ws.max_row
    # print(ws.max_row)
    payoutdata = {}
    if rows_len > 1:
        for i in range(2, rows_len+1):
            # print ws.cell(row=i, column=17).value
            if ws.cell(row=i, column=17).value == today:
                if ws.cell(row=i, column=14).value != "" and ws.cell(row=i, column=14).value is not None:
                    xm = ws.cell(row=i, column=3).value
                    cc = ws.cell(row=i, column=2).value
                    vv = ws.cell(row=i, column=16).value
                    dd = ws.cell(row=i, column=17).value
                    payoutdata['row-'+str(i)] = [xm, cc, vv, dd, str(i)]

    ws = wb[u'setting']
    rows_len = ws.max_row
    setting = {}
    if rows_len > 1:
        mu = ws.cell(row=2, column=1).value
        ms = ws.cell(row=2, column=2).value
        sen = ws.cell(row=2, column=3).value
        rec = []
        cu = ws.cell(row=2, column=5).value
        cs = ws.cell(row=2, column=6).value
        for i in range(2, rows_len+1):
            if ws.cell(row=i, column=4).value is not None:
                receiver = ws.cell(row=i, column=4).value
                rec.append(receiver)
        setting = [mu, ms, sen, cu, cs]
    return payoutdata, setting, rec


def payoutcheck(o, payoutdata, crm_user):
    now_time11 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    print('\t\t'+now_time11+u' 开始')
    # 调用后台静默
    path = os.path.dirname(sys.argv[0])
    os.chdir(path)
    if o == 1:
        driver = webdriver.PhantomJS()
    # 窗口前端操作
    else:
        driver = webdriver.Chrome()
    # driver.maximize_window()
    # 登录操作
    driver.get("http://10.100.20.1:8080/fortune/account/user/main")
    driver.find_element_by_xpath('//*[@id="username"]').send_keys(crm_user[0])
    driver.find_element_by_xpath('//*[@id="password"]').send_keys(crm_user[1])
    driver.find_element_by_xpath('//*[@id="fm1"]/div[3]/a[1]').click()
    # driver.get("http://10.100.20.1:8080/fortune/payoutAudit/auditList")
    name0 = now_time1 + '付款审核情况：'
    for k in payoutdata.keys():
        xm = payoutdata[k][0]
        cc = payoutdata[k][1]
        vv = payoutdata[k][2]
        dd = payoutdata[k][3]
        m = payoutdata[k][4]
        print(str(m), 'xm', xm, 'cc', cc, 'dd', dd, 'vv', vv)
        # 进入卖单界面，获取业务状态及投资编号
        driver.get("http://10.100.20.1:8080/fortune/fortune/invest_list?flowIdType=02000002")
        sleep(0.5)
        flag = 1
        while flag == 1:
            try:
                driver.find_element_by_xpath('//*[@id="lendingNo"]')
                flag = 2
            except:
                sleep(0.5)
        # 填入姓名
        driver.find_element_by_xpath('//*[@id="cnName"]').send_keys(Keys.CONTROL, 'a')
        driver.find_element_by_xpath('//*[@id="cnName"]').send_keys(Keys.BACK_SPACE)
        driver.find_element_by_xpath('//*[@id="cnName"]').send_keys(xm)
        # 填入出借编号
        driver.find_element_by_xpath('//*[@id="lendingNo"]').send_keys(Keys.CONTROL, 'a')
        driver.find_element_by_xpath('//*[@id="lendingNo"]').send_keys(Keys.BACK_SPACE)
        driver.find_element_by_xpath('//*[@id="lendingNo"]').send_keys(cc)
        # 选择全部状态
        driver.find_element_by_xpath('//*[@id="pflowId"]').click()
        sleep(0.5)
        driver.find_element_by_xpath('//*[@id="mainForm"]/div[2]/table/tbody/tr[1]/td/div[1]/label[1]/input').click()
        sleep(0.5)
        compare1 = driver.find_element_by_xpath('//*[@id="mainForm"]/div[3]/p').text
        driver.find_element_by_xpath('//*[@id="query"]').click()
        compare2 = driver.find_element_by_xpath('//*[@id="mainForm"]/div[3]/p').text
        while compare1 == compare2:
            sleep(0.5)
            compare2 = driver.find_element_by_xpath('//*[@id="mainForm"]/div[3]/p').text
        text1 = driver.find_element_by_xpath('//*[@id="ggkj_table"]/tbody/tr[2]/td[5]/a').text
        text2 = driver.find_element_by_xpath('//*[@id="ggkj_table"]/tbody/tr[2]/td[4]').text
        text3 = driver.find_element_by_xpath('//*[@id="ggkj_table"]/tbody/tr[2]/td[12]').text
        flag = 1
        biubiu = 0
        tzbh = ''
        while flag == 1:
            if text1 == xm and text2 == cc and text3 == u'投资结束':
                tzbh = driver.find_element_by_xpath('//*[@id="ggkj_table"]/tbody/tr[2]/td[2]').text
                tzbh = str(tzbh)
                flag = 2
            else:
                sleep(0.5)
                biubiu += 1
                if biubiu > 10:
                    print(str(m) + str(cc) + '失败 ：' + text3)
                    name = str(m) + '\t' + str(cc) + '\t' + '失败 ：' + text3
                    name0 = name0 + '\n' + name
                    break
        if biubiu >= 10:
            continue
        # 进入付款审核界面
        driver.get("http://10.100.20.1:8080/fortune/payoutAudit/auditList")
        sleep(0.5)
        flag = 1
        while flag == 1:
            try:
                driver.find_element_by_xpath('//*[@id="t1"]/div[1]/table/tbody/tr[1]/td/input[1]')
                flag = 2
            except:
                sleep(0.5)
        # 填入投资编号
        driver.find_element_by_xpath('//*[@id="t1"]/div[1]/table/tbody/tr[1]/td/input[1]').send_keys(Keys.CONTROL, 'a')
        driver.find_element_by_xpath('//*[@id="t1"]/div[1]/table/tbody/tr[1]/td/input[1]').send_keys(Keys.BACK_SPACE)
        driver.find_element_by_xpath('//*[@id="t1"]/div[1]/table/tbody/tr[1]/td/input[1]').send_keys(tzbh)
        sel = driver.find_element_by_xpath('//*[@id="recoverWay"]')
        Select(sel).select_by_value('')
        driver.find_element_by_xpath('//*[@id="t1"]/div[1]/table/tbody/tr[1]/td/input[6]').click()
        flag = 1
        diudiu = 0
        while flag == 1:
            try:
                driver.find_element_by_xpath('//*[@id="paybacktable"]/tbody/tr[2]/td[3]')
                flag = 2
            except:
                sleep(0.5)
                diudiu += 1
                if diudiu > 10:
                    print(str(m) + str(cc) + '失败 ：无待审核业务')
                    name = str(m) + '\t' + str(cc) + '\t' + '失败 ：无待审核业务'
                    name0 = name0 + '\n' + name
                    break
        if diudiu >= 10:
            continue
        text1 = driver.find_element_by_xpath('//*[@id="paybacktable"]/tbody/tr[2]/td[4]').text
        text2 = driver.find_element_by_xpath('//*[@id="paybacktable"]/tbody/tr[2]/td[3]').text
        text3 = driver.find_element_by_xpath('//*[@id="paybacktable"]/tbody/tr[2]/td[9]').text
        text3 = float(text3)
        flag = 1
        biubiu = 0
        while flag == 1:
            if text1 == xm and text2 == cc and text3 == float(vv):
                driver.find_element_by_xpath('//*[@id="paybacktable"]/tbody/tr[2]/td[11]/a').click()
                flag = 1
                while flag == 1:
                    try:
                        driver.find_element_by_xpath('//*[@id="memo"]')
                        flag = 2
                    except:
                        sleep(0.5)
                driver.find_element_by_xpath('//*[@id="payoutauditform"]/table/tbody/tr[8]/td/div/input[1]').click()
                print(str(m) + str(cc) + '完成')
                name = str(m) + '\t' + str(cc) + '\t' + '完成'
                name0 = name0 + '\n' + name
                flag = 2
            else:
                sleep(0.5)
                biubiu += 1
                if biubiu > 10:
                    print(str(m) + str(cc) + '失败：付款信息有误')
                    name = str(m) + '\t' + str(cc) + '\t' + '失败：付款信息有误'
                    name0 = name0 + '\n' + name
                    break
        if diudiu >= 10:
            continue
    sleep(0.5)
    driver.quit()
    return name0


if __name__ == '__main__':

    payoutdata0 = get_data()[0]
    crm_user0 = [get_data()[1][3], get_data()[1][4]]
    stmpset0 = [get_data()[1][0], get_data()[1][1], get_data()[1][2], get_data()[2]]
    # print(stmpset0)
    zw01 = ''
    now_time1 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    zw01 = payoutcheck(1, payoutdata0, crm_user0)  # 第一个参数1为后台，其余为前端操作。
    zw0 = zw01
    time.sleep(0.5)
    now_time1 = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    print('\t\t'+now_time1+u'已完成！！！')

    fs0 = u"付款审核PY"
    js0 = u"王上"
    zt0 = u'付款审核完成邮件'
    zw0 = zw0
    smtp(zw0, fs0, js0, zt0, stmpset0)


